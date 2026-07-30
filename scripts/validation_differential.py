#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1", "requests>=2.31"]
# ///
"""Differentially test `gst validate` against GSTN's own offline tool.

Every rule in `spec/gstr1/*.json` is a claim about what the official tool
rejects. This script checks those claims by construction: for each declared
constraint on each field it builds a workbook row that violates it, feeds that
row to BOTH validators, and compares the verdicts.

    uv run scripts/validation_differential.py --app-dir ~/gst-offline-tool/app

The tool's mapping runs server-side, so its verdict is read from the working
file it writes: with exactly one data row in one sheet, an empty section means
the row was rejected. That is a more reliable signal than the error lists
`/addtblfile` returns, which mix row numbers with invoice numbers and leave
some rejections unreported.

Two directions of disagreement mean very different things:

  * The tool rejects and we accept — a MISS. The portal would reject a return
    this implementation called clean. There should be none of these.
  * We reject and the tool accepts — stricter than the reference. Sometimes a
    bug, often deliberate: several of the tool's "accepts" produce malformed
    payloads (a blank taxable value emits `iamt: null`, a bogus invoice type
    drops the `inv_typ` key entirely). Check what the tool actually EMITS
    before treating one of these as a defect here.

Requires the official Returns Offline Tool running locally (`node app.js`).
It is not part of the test suite: it needs a proprietary tool this repository
neither ships nor depends on. See `fixtures/golden/README.md` for the capture
method it builds on.
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
from collections import Counter
from pathlib import Path

import openpyxl
import requests

from _spec import SPEC_DIR, load_specs

REPO = Path(__file__).resolve().parent.parent
ENVELOPE = json.loads((SPEC_DIR / "upload-envelope.json").read_text())
FINANCIAL_YEARS = json.loads((REPO / "spec/masters/financial-years.json").read_text())["years"]
# months.json lists calendar months January-first; rotate to the April-first
# financial-year ordering the tool's UI uses.
_MONTH_NAMES = json.loads((REPO / "spec/masters/months.json").read_text())["names"]
MONTHS = [(name, i + 1) for i, name in enumerate(_MONTH_NAMES)]
MONTHS = MONTHS[3:] + MONTHS[:3]


def share_data(gstin: str, period: str, fy: str, is_tpq: bool = False) -> dict:
    """The context blob `/addtblfile` expects, reconstructed from the UI's."""
    years = []
    for fy_label in FINANCIAL_YEARS:
        # April rollover: January-March belong to the FY's second calendar
        # year. crates/gst-core/src/date.rs `period_from_financial_year` is
        # the Rust counterpart of this rule.
        start = int(fy_label[:4])
        months = [
            {"month": name, "value": f"{m:02d}{start if m >= 4 else start + 1}"}
            for name, m in MONTHS
        ]
        years.append({"year": fy_label, "months": months})
    current = next(y for y in years if y["year"] == fy)["months"]
    month_name = next(n for n, m in MONTHS if m == int(period[:2]))
    return {
        "dashBoardDt": {"form": "GSTR1", "gstin": gstin, "fp": period},
        "yearsList": years,
        "curFyMonths": current,
        "monthSelected": {"month": month_name, "value": period},
        "yearSelected": {"year": fy, "months": current},
        "isSezTaxpayer": False,
        "isUploadImport": False,
        "isTPQ": is_tpq,
        "disableHSNRestrictions": False,
        "newHSNStartDateConstant": "052021",
        "R1_NEW_ECO_STRT_PRD": "012024",
        "R1_NEW_ECOA_STRT_PRD": "022024",
    }


def violations(field: dict) -> list[tuple[str, str]]:
    """Values the spec claims must be rejected, one per declared constraint."""
    out: list[tuple[str, str]] = []
    ty = field.get("type")
    if field.get("required") and not field.get("default"):
        out.append(("required_blank", ""))
    if field.get("enum"):
        out.append(("enum_bogus", "ZZBOGUS"))
    if field.get("enum_ref"):
        out.append(("enum_ref_bogus", "9999" if "tax-rates" in field["enum_ref"] else "ZZBOGUS"))
    if ty == "state_code":
        out.append(("pos_out_of_range", "99-Nowhere"))
    if ty == "gstin" and not field.get("must_be_empty"):
        out.append(("gstin_bad_checkdigit", "27AAPFU0939F1ZA"))
        out.append(("gstin_too_short", "27AAPFU0939F1Z"))
    if ty == "date":
        out.append(("date_not_a_calendar_date", "32-Jan-2018"))
        out.append(("date_before_gst", "30-Jun-2017"))
    pattern = field.get("pattern")
    if pattern and ty == "decimal":
        # Only meaningful where the reference does NOT round first.
        if not field.get("round_to"):
            out.append(("decimal_too_many_places", "1.234"))
        if not pattern.startswith("^-?"):
            out.append(("decimal_negative_not_allowed", "-100"))
    if pattern and ty == "text":
        out.append(("text_illegal_characters", "A$B#C"))
        if field.get("max_length"):
            out.append(("text_too_long", "A" * (int(field["max_length"]) + 4)))
    if field.get("must_be_empty"):
        out.append(("must_be_empty_violated", "12AJIPA1572E1C7"))
    return out


def base_rows(workbook: Path, specs: dict[str, dict]) -> dict[str, dict]:
    excel = {s["source"]["excel"]["sheet"]: s["source"]["excel"] for s in specs.values()}
    wb = openpyxl.load_workbook(workbook)
    out = {}
    for ws in wb.worksheets:
        src = excel.get(ws.title)
        if src is None:
            continue
        header = [c.value for c in ws[src["header_row"]] if c.value is not None]
        row = [c.value for c in ws[src["first_data_row"]]][: len(header)]
        out[ws.title] = {
            "header": header,
            "row": ["" if v is None else v for v in row],
        }
    return out


def in_period(spec: dict, period: str) -> bool:
    """Whether a section is filed for this period at all. A section outside its
    window would be probed against a tool that never reads that sheet, and every
    case would read as a disagreement."""
    def yyyymm(p: str) -> int:
        return int(p[2:]) * 100 + int(p[:2])

    now = yyyymm(period)
    start, end = spec.get("active_from_period"), spec.get("active_until_period")
    if start and now < yyyymm(start):
        return False
    if end and now >= yyyymm(end):
        return False
    return True


def build_cases(specs: dict[str, dict], bases: dict[str, dict], period: str) -> list[dict]:
    cases = []
    for section, spec in specs.items():
        excel = spec["source"]["excel"]
        sheet = excel["sheet"]
        if sheet not in bases or not in_period(spec, period):
            continue
        header, row = bases[sheet]["header"], bases[sheet]["row"]
        layout = {
            "sheet": sheet, "header_row": excel["header_row"],
            "first_data_row": excel["first_data_row"],
        }
        cases.append({
            "section": section, "case": "control", "field": None,
            "value": None, "header": header, "row": row, **layout,
        })
        for field in spec["fields"]:
            if field["column"] not in header:
                continue
            for label, value in violations(field):
                mutated = list(row)
                mutated[header.index(field["column"])] = value
                cases.append({
                    "section": section,
                    "case": f"{field['id']}.{label}", "field": field["id"],
                    "value": value, "header": header, "row": mutated, **layout,
                })
    return cases


def record_paths(envelope: dict) -> dict[str, tuple[list, list]]:
    """Where each section's records sit inside the tool's working file,
    derived from the upload-envelope spec's `keys`: section code ->
    (paths, fallback paths), each path a list of JSON keys. Sections not
    mapped here are the plain `section:` keys — a bare array under their
    own top-level key."""
    paths: dict[str, tuple[list, list]] = {}

    def add(section: str, path: list[str], fallback: bool = False) -> None:
        entry = paths.setdefault(section, ([], []))
        entry[1 if fallback else 0].append(path)

    for entry in envelope["keys"]:
        src = entry.get("from", "")
        if src.startswith("wrapped:"):
            add(src.removeprefix("wrapped:"), [entry["key"], entry["wrapper"]])
        elif src == "object":
            for member in entry["members"]:
                add(member["from"].removeprefix("section:"), [entry["key"], member["key"]])
    combined = envelope["hsn_before_bifurcation"]["wrapper"]
    for member in envelope["hsn_from_bifurcation"]["members"]:
        section = member["from"].removeprefix("section:")
        add(section, ["hsn", member["key"]])
        add(section, ["hsn", combined], fallback=True)
    # The pre-bifurcation HSN table is not a `section:` key in the envelope —
    # `hsn` is built by its own branch — so it needs mapping by hand or its
    # records are looked for at the top level, where an object sits rather than
    # an array, and every row reads as a rejection.
    add("hsn", ["hsn", combined])
    for member in envelope["hsn_from_bifurcation"]["members"]:
        add("hsn", ["hsn", member["key"]], fallback=True)
    return paths


RECORD_PATHS = record_paths(ENVELOPE)


def dig(work: dict, path: list[str]) -> list:
    for key in path:
        work = work.get(key, {}) if isinstance(work, dict) else {}
    return work if isinstance(work, list) else []


class Tool:
    """The official offline tool, driven over its own HTTP endpoints."""

    def __init__(self, app_dir: Path, gstin: str, period: str, fy: str, month: str, port: int,
                 is_tpq: bool = False):
        self.base = f"http://localhost:{port}"
        self.gstin, self.period, self.fy, self.month = gstin, period, fy, month
        self.work_dir = app_dir / "public/userData" / gstin / "GSTR1" / fy / month
        self.work_file = self.work_dir / f"GSTR1_{gstin}_{fy}_{month}.json"
        self.share = json.dumps(share_data(gstin, period, fy, is_tpq))

    def records(self, work: dict, section: str) -> int:
        primary, fallback = RECORD_PATHS.get(section, ([[section]], []))
        count = sum(len(dig(work, path)) for path in primary)
        if count == 0 and fallback:
            count = sum(len(dig(work, path)) for path in fallback)
        return count

    def seed(self, workbook: Path) -> dict:
        """Clean-slate import of one workbook into the tool's working store:
        `/addtblfile` then `/addmltpldata`, the same two steps the golden
        captures document. Returns the import response (per-section rejection
        lists plus `cache_key`); raises on a hung or non-JSON import."""
        shutil.rmtree(self.work_dir, ignore_errors=True)
        with workbook.open("rb") as fh:
            resp = requests.post(
                f"{self.base}/addtblfile",
                files={"file": (workbook.name, fh)},
                data={"shareData": self.share},
                timeout=30,
            ).json()
        requests.post(
            f"{self.base}/addmltpldata",
            json={
                "gstin": self.gstin, "form": "GSTR1", "fy": self.fy, "month": self.month,
                "fp": self.period, "gt": "", "cur_gt": "", "type": "",
                "tbl_data": {"cache_key": resp.get("cache_key")},
            },
            timeout=60,
        )
        return resp

    def verdict(self, workbook: Path, section: str) -> tuple[str, str]:
        """'accept', 'reject', or 'crash' — the tool has unhandled paths that
        hang the request rather than returning an error, and scoring those as
        acceptances would hide a defect that is worse than either."""
        try:
            resp = self.seed(workbook)
        except (requests.RequestException, ValueError):
            return "crash", "import hung or returned nothing"
        fired = ",".join(k for k, v in resp.items() if k != "cache_key" and v)
        if not self.work_file.exists():
            return "reject", fired or "no-working-file"
        kept = self.records(json.loads(self.work_file.read_text()), section)
        return ("reject" if kept == 0 else "accept"), fired


def ours_rejects(binary: Path, workbook: Path, section: str, gstin: str, period: str):
    out = subprocess.run(
        [str(binary), "validate", str(workbook), "--section", section,
         "--gstin", gstin, "--period", period, "--format", "json"],
        capture_output=True, text=True, cwd=REPO,
    )
    try:
        findings = json.loads(out.stdout)
    except json.JSONDecodeError:
        return True, "unparseable:" + (out.stderr or out.stdout).strip()[:80]
    errors = [f for f in findings if f.get("severity") == "error"]
    return bool(errors), ";".join(sorted({f.get("rule") or f.get("column") or "?" for f in errors}))


def write_case(case: dict, path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = case["sheet"]
    ws.cell(1, 1, f"Summary for {case['sheet']}")
    for i, head in enumerate(case["header"], 1):
        ws.cell(case["header_row"], i, head)
    for i, value in enumerate(case["row"], 1):
        if value != "":
            ws.cell(case["first_data_row"], i, value)
    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--app-dir", required=True, type=Path,
                    help="the running offline tool's app directory")
    ap.add_argument("--workbook", type=Path,
                    default=REPO / "fixtures/gstr1/demo-workbook.xlsx",
                    help="workbook supplying a valid base row per section")
    ap.add_argument("--gstin", default="27AAPFU0939F1ZV")
    ap.add_argument("--period", default="062025")
    ap.add_argument("--fy", default="2025-26")
    ap.add_argument("--month", default="June")
    ap.add_argument("--port", type=int, default=3010)
    ap.add_argument("--binary", type=Path, default=REPO / "target/debug/gst")
    ap.add_argument("--out", type=Path, help="write the full result table as JSON")
    args = ap.parse_args()

    if not args.binary.exists():
        sys.exit(f"{args.binary} not found — run `cargo build -p gst-cli` first")
    try:
        requests.get(f"http://localhost:{args.port}/health", timeout=5)
    except requests.RequestException:
        sys.exit(f"no offline tool on port {args.port} — start it with `node app.js`")

    specs = load_specs()
    cases = build_cases(specs, base_rows(args.workbook, specs), args.period)
    tool = Tool(args.app_dir, args.gstin, args.period, args.fy, args.month, args.port)
    scratch = Path(args.out).parent if args.out else REPO
    workbook = scratch / ".differential-case.xlsx"

    results = []
    print(f"{len(cases)} cases derived from the spec", flush=True)
    for i, case in enumerate(cases, 1):
        write_case(case, workbook)
        tool_verdict, fired = tool.verdict(workbook, case["section"])
        mine, why = ours_rejects(args.binary, workbook, case["section"], args.gstin, args.period)
        results.append({
            "section": case["section"], "case": case["case"], "field": case["field"],
            "value": case["value"], "tool": tool_verdict, "ours_rejects": mine,
            "tool_note": fired, "ours_note": why,
        })
        if i % 40 == 0:
            print(f"  {i}/{len(cases)}", flush=True)
    workbook.unlink(missing_ok=True)

    if args.out:
        args.out.write_text(json.dumps(results, indent=1))

    misses = [r for r in results if r["tool"] == "reject" and not r["ours_rejects"]]
    stricter = [r for r in results if r["tool"] == "accept" and r["ours_rejects"]]
    crashes = [r for r in results if r["tool"] == "crash"]
    agree = [r for r in results if (r["tool"] == "reject") == r["ours_rejects"]
             and r["tool"] != "crash"]
    controls = [r for r in results if r["case"] == "control"
                and (r["tool"] != "accept" or r["ours_rejects"])]

    print(f"\n{len(agree)}/{len(results)} agree")
    print(f"  MISSES (tool rejects, we accept): {len(misses)}")
    for r in misses:
        print(f"     {r['section']:9} {r['case']}")
    print(f"  stricter than the reference: {len(stricter)}")
    for kind, n in Counter(r["case"].split(".", 1)[1] for r in stricter).most_common():
        print(f"     {n:3}  {kind}")
    print(f"  tool CRASHED (import hangs, no verdict): {len(crashes)}")
    for r in crashes:
        print(f"     {r['section']:9} {r['case']}")
    if controls:
        print(f"  WARNING: {len(controls)} control rows did not pass cleanly — the base "
              f"workbook or shareData is wrong, so the run is not trustworthy")
        for r in controls:
            print(f"     {r['section']}: tool={r['tool']} ours_rejects={r['ours_rejects']}")
    return 1 if misses or controls else 0


if __name__ == "__main__":
    sys.exit(main())
